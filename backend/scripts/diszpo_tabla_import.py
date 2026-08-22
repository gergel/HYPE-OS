"""A HYPE 2026 Google Sheet táblázat átvétele - MINDEN munkalappal és SZÍNNEL.

    python scripts/diszpo_tabla_import.py                     # próba, nem ír
    python scripts/diszpo_tabla_import.py --vegrehajt         # élesben
    python scripts/diszpo_tabla_import.py --fajl hype.xlsx    # helyi fájlból

A táblázat MEGOSZTOTT linkkel érhető el, tehát nem kell hozzá Google-fiók: az
`export?format=xlsx` végpont adja vissza az egész munkafüzetet - és ami a
lényeg, a CELLÁK SZÍNÉVEL együtt. A CSV-export ezt nem tudná, pedig itt a szín
maga az adat: az mondja meg, ki melyik nap dolgozott.

ÚJRAFUTTATHATÓ. A második átvételnél (amikor teljesen átálltok a rendszerre)
ugyanígy lefuttatható: munkalaponként cseréli a tartalmat, tehát nem duplázza
a sorokat. Ami viszont NÁLUNK készült és a Sheetben nincs benne, azt a csere
ELDOBJA - ezért a próba-futás kiírja, mennyi cella érintett, mielőtt bármit
tenne.

AZ OSZLOP-EMBER KÖTÉS a fejlécben álló keresztnévre megy ("GERI" ->
Gergely...). Ahol a név nem azonosít EGYÉRTELMŰEN egy belsős munkatársat, ott
a kötés ÜRESEN marad, és a szkript kiírja - két azonos keresztnév között egy
szkript nem tud dönteni, és a rossz kötés csendben más ember napjait
számolná. A felületen utólag megadható.
"""

from __future__ import annotations

import argparse
import io
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    import openpyxl  # noqa: E402
except ModuleNotFoundError as exc:  # pragma: no cover - üzemeltetési segítség
    # Az openpyxl csak ennek a szkriptnek kell, és a régi image-ekben még nincs
    # benne. A puszta ModuleNotFoundError itt nem mond semmit arról, mit kell
    # tenni - ezért kiírjuk.
    raise SystemExit(
        "Hiányzik az openpyxl csomag, enélkül nem tudjuk beolvasni a táblázatot.\n"
        "  - ha most, egyszer futtatnád:  pip install openpyxl\n"
        "  - tartósan: a requirements.txt-ben már benne van, tehát a következő\n"
        "    deploy után magától meglesz."
    ) from exc

import requests  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.models.diszpo_tabla import (  # noqa: E402
    SZIN_FEHER,
    SZIN_KEK,
    SZIN_PIROS,
    SZIN_SZURKE,
    SZIN_ZOLD,
    DiszpoCella,
    DiszpoMunkalap,
    DiszpoOszlop,
    DiszpoSor,
)
from app.models.employee import Employee, EmployeeType  # noqa: E402
from app.services.hu_szoveg import ekezet_nelkul  # noqa: E402

#: A megosztott munkafüzet azonosítója (a link /d/ és /edit közti része).
TABLAZAT_ID = "1Xflz0Ig3z7bgoN5hspeRIICHbY4P8XIcxCQ3fQPJcnA"

#: A Sheet hexakódjai -> a mi elnevezett színeink. Ami nincs a listán, az
#: nálunk szín nélkül marad: a jelentés nélküli szín csak zavarna a
#: számolásnál (lásd models/diszpo_tabla.py).
SZIN_TERKEP: dict[str, str] = {
    "FF00FF00": SZIN_ZOLD,
    "FFFF0000": SZIN_PIROS,
    "FFFFFFFF": SZIN_FEHER,
    "FFB7B7B7": SZIN_SZURKE,
    "FF4A86E8": SZIN_KEK,
}

#: A belsős munkalapon a fejléc két sor: a felső a SZEKCIÓ ("CAMERA CREW"), az
#: alsó a nevek. A többi munkalapon egy.
KETSOROS_FEJLECU: frozenset[str] = frozenset({"BELSŐS DISZPÓSTÁBLA"})

#: Hónap-elválasztó sorok felismerése ("❄️ JANUÁR ❄️").
HONAPOK: tuple[str, ...] = (
    "januar", "februar", "marcius", "aprilis", "majus", "junius",
    "julius", "augusztus", "szeptember", "oktober", "november", "december",
)


def letoltes(tablazat_id: str) -> bytes:
    cim = f"https://docs.google.com/spreadsheets/d/{tablazat_id}/export?format=xlsx"
    valasz = requests.get(cim, timeout=120)
    valasz.raise_for_status()
    return valasz.content


def cella_szine(cella) -> str | None:
    kitoltes = cella.fill
    if kitoltes is None or kitoltes.fill_type is None:
        return None
    rgb = getattr(kitoltes.fgColor, "rgb", None)
    return SZIN_TERKEP.get(rgb) if isinstance(rgb, str) else None


def cella_erteke(ertek) -> str | None:
    """A cella szövege. A dátum/idő értékeket olvasható alakra hozzuk - a
    táblázatban is így látszanak, és egy "2026-08-16 00:00:00" a beosztásban
    csak zavar."""
    if ertek is None:
        return None
    if isinstance(ertek, datetime):
        return ertek.strftime("%Y.%m.%d.") if (ertek.hour, ertek.minute) == (0, 0) else ertek.strftime("%H:%M")
    if isinstance(ertek, date):
        return ertek.strftime("%Y.%m.%d.")
    if isinstance(ertek, float) and ertek.is_integer():
        return str(int(ertek))
    szoveg = str(ertek).strip()
    return szoveg or None


def elvalaszto_sor(ertekek: list[str | None]) -> bool:
    szoveg = ekezet_nelkul(" ".join(e for e in ertekek if e))
    return any(h in szoveg for h in HONAPOK) and len([e for e in ertekek if e]) <= 3


def tartalom_hatara(ws) -> tuple[int, int]:
    """A ténylegesen HASZNÁLT terület - a Sheet elvi 1000 sora helyett.

    Nem csak az értéket nézzük, hanem a SZÍNT is: a belsős táblán rengeteg a
    kitöltött, de szöveg nélküli cella (a piros szabadnapok), és azok nélkül
    a beosztás fele eltűnne."""
    max_sor = max_oszlop = 0
    for sor in ws.iter_rows():
        for cella in sor:
            if cella.value is not None or cella_szine(cella) is not None:
                max_sor = max(max_sor, cella.row)
                max_oszlop = max(max_oszlop, cella.column)
    return max_sor, max_oszlop


def belsos_nevterkep(db) -> dict[str, list[Employee]]:
    """Keresztnév (ékezet nélkül, kisbetűvel) -> munkatársak.

    Nem csak a belsősöket vesszük: a táblázat oszlopai közt vágó és kreatív is
    van. Az "aki nálunk dolgozik" a szűrő, nem a típus."""
    terkep: dict[str, list[Employee]] = {}
    for emb in db.scalars(select(Employee).where(Employee.is_active.is_not(False))).all():
        for resz in (emb.full_name or "").split():
            terkep.setdefault(ekezet_nelkul(resz), []).append(emb)
    return terkep


def oszlop_embere(cimke: str | None, terkep: dict[str, list[Employee]]) -> tuple[Employee | None, str | None]:
    """(munkatárs, üzenet) - az oszlop fejlécében álló név alapján.

    Csak EGYÉRTELMŰ találatot fogadunk el: ha a keresztnévre két ember is
    illik, a kötés üresen marad. Egy rossz kötés csendben MÁS ember napjait
    számolná, és ez a szám a projektek önköltségébe megy - ott a hallgatólagos
    tévedés a legdrágább."""
    if not cimke:
        return None, None
    kulcs = ekezet_nelkul(cimke.strip())
    jeloltek = terkep.get(kulcs, [])
    if len(jeloltek) == 1:
        return jeloltek[0], None
    if not jeloltek:
        return None, f"„{cimke}” - nincs ilyen nevű munkatárs, a kötés üres marad"
    nevek = ", ".join(e.full_name for e in jeloltek[:4])
    return None, f"„{cimke}” - többen is illenek rá ({nevek}), a kötést kézzel kell megadni"


def munkalap_atvetele(db, ws, sorrend: int, vegrehajt: bool) -> dict:
    max_sor, max_oszlop = tartalom_hatara(ws)
    fejlec_sorok = 2 if ws.title in KETSOROS_FEJLECU else 1

    meglevo = db.scalar(select(DiszpoMunkalap).where(DiszpoMunkalap.nev == ws.title))
    regi_cellak = (
        db.scalar(select(DiszpoCella.id).where(DiszpoCella.munkalap_id == meglevo.id).limit(1)) is not None
        if meglevo
        else False
    )

    # A CELLÁK
    cellak: list[tuple[int, int, str | None, str | None]] = []
    for r in range(1, max_sor + 1):
        for c in range(1, max_oszlop + 1):
            cella = ws.cell(r, c)
            ertek = cella_erteke(cella.value)
            szin = cella_szine(cella)
            if ertek is None and szin is None:
                continue
            cellak.append((r - 1, c - 1, ertek, szin))

    # AZ OSZLOPOK: a fejléc alsó sora a címke, a felső a szekció - a szekció a
    # Sheetben csak az első oszlopánál áll ott, tehát jobbra "átfolyik".
    csoport_sor = 1 if fejlec_sorok == 2 else 0
    nevsor = fejlec_sorok - 1
    terkep = belsos_nevterkep(db)
    oszlopok: list[dict] = []
    uzenetek: list[str] = []
    aktualis_csoport: str | None = None
    for c in range(1, max_oszlop + 1):
        if fejlec_sorok == 2:
            csoport_ertek = cella_erteke(ws.cell(1, c).value)
            if csoport_ertek:
                aktualis_csoport = csoport_ertek
        cimke = cella_erteke(ws.cell(nevsor + 1, c).value)
        ember, uzenet = oszlop_embere(cimke, terkep) if fejlec_sorok == 2 else (None, None)
        if uzenet:
            uzenetek.append(uzenet)
        oszlopok.append(
            {
                "idx": c - 1,
                "cimke": cimke,
                "csoport": aktualis_csoport if fejlec_sorok == 2 else None,
                "employee_id": ember.id if ember else None,
            }
        )

    # A SOROK: a dátum TOVÁBBVITELE a lényeg - egy naphoz két diszpó is
    # tartozhat, és a második sor dátum-mezője üres.
    sorok: list[dict] = []
    utolso_datum: date | None = None
    for r in range(1, max_sor + 1):
        nyers_datum = ws.cell(r, 1).value if max_oszlop >= 1 else None
        if isinstance(nyers_datum, datetime):
            utolso_datum = nyers_datum.date()
        elif isinstance(nyers_datum, date):
            utolso_datum = nyers_datum
        ertekek = [cella_erteke(ws.cell(r, c).value) for c in range(1, min(max_oszlop, 8) + 1)]
        elvalaszto = elvalaszto_sor(ertekek)
        if elvalaszto:
            utolso_datum = None
        nap = cella_erteke(ws.cell(r, 2).value) if max_oszlop >= 2 else None
        diszposzam = cella_erteke(ws.cell(r, 3).value) if max_oszlop >= 3 else None
        sorok.append(
            {
                "idx": r - 1,
                "datum": utolso_datum if r > fejlec_sorok and not elvalaszto else None,
                "nap": nap if isinstance(nap, str) and len(nap) < 20 else None,
                "diszposzam": int(diszposzam) if (diszposzam or "").isdigit() else None,
                "elvalaszto": elvalaszto,
            }
        )

    osszegzes = {
        "munkalap": ws.title,
        "sorok": len(sorok),
        "oszlopok": len(oszlopok),
        "cellak": len(cellak),
        "szines": sum(1 for *_, szin in cellak if szin),
        "emberhez_kotve": sum(1 for o in oszlopok if o["employee_id"]),
        "uzenetek": uzenetek,
        "felulir": regi_cellak,
    }
    if not vegrehajt:
        return osszegzes

    if meglevo is None:
        meglevo = DiszpoMunkalap(nev=ws.title)
        db.add(meglevo)
    meglevo.sorrend = sorrend
    meglevo.sor_szam = len(sorok)
    meglevo.oszlop_szam = len(oszlopok)
    meglevo.fejlec_sorok = fejlec_sorok
    db.flush()

    # CSERE, nem összefésülés: a Sheet az igazság ennél az átvételnél.
    for tabla in (DiszpoCella, DiszpoSor, DiszpoOszlop):
        db.query(tabla).filter(tabla.munkalap_id == meglevo.id).delete(synchronize_session=False)
    db.flush()

    db.bulk_save_objects([DiszpoOszlop(munkalap_id=meglevo.id, **o) for o in oszlopok])
    db.bulk_save_objects([DiszpoSor(munkalap_id=meglevo.id, **s) for s in sorok])
    db.bulk_save_objects(
        [
            DiszpoCella(munkalap_id=meglevo.id, sor_idx=r, oszlop_idx=c, ertek=e, szin=sz)
            for r, c, e, sz in cellak
        ]
    )
    db.commit()
    return osszegzes


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--vegrehajt", action="store_true", help="Ténylegesen írja az adatbázist.")
    parser.add_argument("--fajl", help="Helyi .xlsx (letöltés helyett).")
    parser.add_argument("--tablazat-id", default=TABLAZAT_ID)
    parser.add_argument("--munkalap", action="append", help="Csak ezt a munkalapot (többször megadható).")
    args = parser.parse_args()

    if args.fajl:
        adat = Path(args.fajl).read_bytes()
        print(f"Helyi fájl: {args.fajl}")
    else:
        print("Letöltés a Google Sheetsből…")
        adat = letoltes(args.tablazat_id)
    wb = openpyxl.load_workbook(io.BytesIO(adat))

    db = SessionLocal()
    try:
        print(f"\n{'PRÓBA (nem ír)' if not args.vegrehajt else 'VÉGREHAJTÁS'}\n")
        osszes_uzenet: list[str] = []
        for sorrend, nev in enumerate(wb.sheetnames):
            if args.munkalap and nev not in args.munkalap:
                continue
            o = munkalap_atvetele(db, wb[nev], sorrend, args.vegrehajt)
            felulir = " (a meglévő tartalmat CSERÉLI)" if o["felulir"] else ""
            print(
                f"  {o['munkalap']:<24} {o['sorok']:>4} sor x {o['oszlopok']:>3} oszlop, "
                f"{o['cellak']:>6} cella ({o['szines']} színes), "
                f"{o['emberhez_kotve']} oszlop emberhez kötve{felulir}"
            )
            osszes_uzenet.extend(o["uzenetek"])

        if osszes_uzenet:
            print("\nAMIT KÉZZEL KELL RENDEZNI (az oszlop-ember kötés üresen maradt):")
            for u in dict.fromkeys(osszes_uzenet):
                print("  -", u)
            print("\n  Ezek nélkül az érintett oszlop napjai nem számítanak bele a")
            print("  munkanap-számlálásba. A kötés a felületen megadható.")

        if not args.vegrehajt:
            print("\nEz csak PRÓBA volt. Éles futtatás: --vegrehajt")
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
