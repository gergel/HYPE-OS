"""A HYPE 2026 Google Sheet szinkronja - MINDEN munkalappal és SZÍNNEL.

A scripts/diszpo_tabla_import.py magja szolgáltatásba emelve, hogy a
szinkron a FELÜLETRŐL is indítható legyen (lásd routes/diszpo_tabla.py
"sheet-sync" végpontja): a felhasználó egy gombbal ellenőrizheti/átveheti,
hogy a rendszer pontosan azt mutassa, ami a Google Táblázatban van.

A táblázat MEGOSZTOTT linkkel érhető el, tehát nem kell hozzá Google-fiók: az
`export?format=xlsx` végpont adja vissza az egész munkafüzetet - és ami a
lényeg, a CELLÁK SZÍNÉVEL együtt (a CSV-export ezt nem tudná, pedig itt a
szín maga az adat: az mondja meg, ki melyik nap dolgozott).

ÚJRAFUTTATHATÓ: munkalaponként CSERÉLI a tartalmat (a Sheet az igazság),
tehát nem duplikál. A kézzel beállított OSZLOP-EMBER kötések viszont
megmaradnak: cserénél az azonos feliratú (cimke) oszlop korábbi kötését
visszaörököljük, ha az automatikus névre-kötés nem talál egyértelmű embert -
enélkül minden szinkron eldobná, amit a felületen egyszer már kézzel
rendbe tettek."""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import requests
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.diszpo_tabla import (
    SZIN_FEHER,
    SZIN_KEK,
    SZIN_PIROS,
    SZIN_SZURKE,
    SZIN_ZOLD,
    DiszpoCella,
    DiszpoMunkalap,
    DiszpoOszlop,
    DiszpoSor,
)
from app.models.employee import Employee
from app.services.hu_szoveg import ekezet_nelkul

#: A megosztott munkafüzet azonosítója (a link /d/ és /edit közti része).
TABLAZAT_ID = "1Xflz0Ig3z7bgoN5hspeRIICHbY4P8XIcxCQ3fQPJcnA"

#: A Sheet hexakódjai -> a mi elnevezett színeink. Ami nincs a listán, az
#: nálunk szín nélkül marad: a jelentés nélküli szín csak zavarna a
#: számolásnál (lásd models/diszpo_tabla.py).
SZIN_TERKEP: dict[str, str] = {
    "FF00FF00": SZIN_ZOLD,
    "FFFF0000": SZIN_PIROS,
    "FFFFFFFF": SZIN_FEHER,
    "FFB7B7B7": SZIN_SZURKE,
    "FF4A86E8": SZIN_KEK,
}

#: A belsős munkalapon a fejléc két sor: a felső a SZEKCIÓ ("CAMERA CREW"), az
#: alsó a nevek. A többi munkalapon egy.
KETSOROS_FEJLECU: frozenset[str] = frozenset({"BELSŐS DISZPÓSTÁBLA"})

#: Hónap-elválasztó sorok felismerése ("❄️ JANUÁR ❄️").
HONAPOK: tuple[str, ...] = (
    "januar", "februar", "marcius", "aprilis", "majus", "junius",
    "julius", "augusztus", "szeptember", "oktober", "november", "december",
)


def letoltes(tablazat_id: str = TABLAZAT_ID) -> bytes:
    cim = f"https://docs.google.com/spreadsheets/d/{tablazat_id}/export?format=xlsx"
    valasz = requests.get(cim, timeout=120)
    valasz.raise_for_status()
    return valasz.content


def cella_szine(cella) -> str | None:
    kitoltes = cella.fill
    if kitoltes is None or kitoltes.fill_type is None:
        return None
    rgb = getattr(kitoltes.fgColor, "rgb", None)
    return SZIN_TERKEP.get(rgb) if isinstance(rgb, str) else None


def cella_erteke(ertek) -> str | None:
    """A cella szövege. A dátum/idő értékeket olvasható alakra hozzuk - a
    táblázatban is így látszanak, és egy "2026-08-16 00:00:00" a beosztásban
    csak zavar."""
    if ertek is None:
        return None
    if isinstance(ertek, datetime):
        return ertek.strftime("%Y.%m.%d.") if (ertek.hour, ertek.minute) == (0, 0) else ertek.strftime("%H:%M")
    if isinstance(ertek, date):
        return ertek.strftime("%Y.%m.%d.")
    if isinstance(ertek, float) and ertek.is_integer():
        return str(int(ertek))
    szoveg = str(ertek).strip()
    return szoveg or None


def elvalaszto_sor(ertekek: list[str | None]) -> bool:
    szoveg = ekezet_nelkul(" ".join(e for e in ertekek if e))
    return any(h in szoveg for h in HONAPOK) and len([e for e in ertekek if e]) <= 3


def tartalom_hatara(ws) -> tuple[int, int]:
    """A ténylegesen HASZNÁLT terület - a Sheet elvi 1000 sora helyett.

    Nem csak az értéket nézzük, hanem a SZÍNT is: a belsős táblán rengeteg a
    kitöltött, de szöveg nélküli cella (a piros szabadnapok), és azok nélkül
    a beosztás fele eltűnne."""
    max_sor = max_oszlop = 0
    for sor in ws.iter_rows():
        for cella in sor:
            if cella.value is not None or cella_szine(cella) is not None:
                max_sor = max(max_sor, cella.row)
                max_oszlop = max(max_oszlop, cella.column)
    return max_sor, max_oszlop


def belsos_nevterkep(db: Session) -> dict[str, list[Employee]]:
    """Keresztnév (ékezet nélkül, kisbetűvel) -> munkatársak."""
    terkep: dict[str, list[Employee]] = {}
    for emb in db.scalars(select(Employee).where(Employee.is_active.is_not(False))).all():
        for resz in (emb.full_name or "").split():
            terkep.setdefault(ekezet_nelkul(resz), []).append(emb)
    return terkep


def oszlop_embere(cimke: str | None, terkep: dict[str, list[Employee]]) -> tuple[Employee | None, str | None]:
    """(munkatárs, üzenet) - az oszlop fejlécében álló név alapján.

    Csak EGYÉRTELMŰ találatot fogadunk el: ha a keresztnévre két ember is
    illik, a kötés üresen marad. Egy rossz kötés csendben MÁS ember napjait
    számolná, és ez a szám a projektek önköltségébe megy."""
    if not cimke:
        return None, None
    kulcs = ekezet_nelkul(cimke.strip())
    jeloltek = terkep.get(kulcs, [])
    if len(jeloltek) == 1:
        return jeloltek[0], None
    if not jeloltek:
        return None, f"„{cimke}” - nincs ilyen nevű munkatárs, a kötés üres marad"
    nevek = ", ".join(e.full_name for e in jeloltek[:4])
    return None, f"„{cimke}” - többen is illenek rá ({nevek}), a kötést kézzel kell megadni"


#: Ennél több felső sort akkor sem fagyasztunk be, ha a dátumok furcsán messze
#: kezdődnek - egy fél képernyőnyi rögzített fejléc használhatatlan lenne.
MAX_FEJLEC_SOROK = 20


def fejlec_sorok_szama(ws, max_sor: int) -> int:
    """Hány felső sor a FEJLÉC: minden, ami az első DÁTUMOS sor fölött áll.

    A külsős táblán a dátumok előtt egy 12 soros blokk van (jelmagyarázat +
    a nevek sora) - a felhasználó kérése, hogy görgetésnél ez az EGÉSZ maradjon
    a helyén, hogy látszódjon, kinek az oszlopában jár az ember. A belsősön a
    dátumok a 3. sorban kezdődnek, ott ez ugyanaz a 2 fejléc-sor, mint eddig.
    Dátum nélküli munkalapon (Autók, AnyDesk...) marad a régi szabály."""
    for r in range(1, max_sor + 1):
        if isinstance(ws.cell(r, 1).value, (datetime, date)):
            fejlec = min(max(r - 1, 1), MAX_FEJLEC_SOROK)
            # A belsős táblán MINDIG legalább két sor rögzül (szekciók + a
            # nevek sora): a felhasználó kérése, hogy görgetés közben is
            # látsszon, kihez ír be - akkor is, ha a Sheetben a dátumok
            # feljebb csúsznának.
            if ws.title in KETSOROS_FEJLECU:
                fejlec = max(fejlec, 2)
            return fejlec
    return 2 if ws.title in KETSOROS_FEJLECU else 1


def nevsor_indexe(ws, fejlec_sorok: int, max_oszlop: int) -> int:
    """A fejléc-blokkon belül az OSZLOPCÍMKÉK (nevek) sora - 1-alapú.

    Az a fejléc-sor, amelyikben a legtöbb kitöltött cella áll a D oszloptól:
    a külsős táblán ez a nevek sora (az A-C a dátum/nap/diszpószám blokk, a
    többi fejléc-sor jelmagyarázat), a belsősön a 2. sor. Korábban a fejléc
    utolsó sorát vettük, ami a külsősön üres volt - így az oszlopoknak nem
    volt címkéjük, és a kézi oszlop-ember kötést sem lehetett a cimkén át
    megőrizni a szinkronok közt."""
    legjobb, legtobb = fejlec_sorok, 0
    for r in range(1, fejlec_sorok + 1):
        kitoltott = sum(1 for c in range(4, max_oszlop + 1) if cella_erteke(ws.cell(r, c).value))
        if kitoltott > legtobb:
            legtobb, legjobb = kitoltott, r
    return legjobb


def munkalap_atvetele(db: Session, ws, sorrend: int, vegrehajt: bool) -> dict:
    max_sor, max_oszlop = tartalom_hatara(ws)
    fejlec_sorok = fejlec_sorok_szama(ws, max_sor)
    ketsoros = ws.title in KETSOROS_FEJLECU

    meglevo = db.scalar(select(DiszpoMunkalap).where(DiszpoMunkalap.nev == ws.title))
    regi_cellak = (
        db.scalar(select(DiszpoCella.id).where(DiszpoCella.munkalap_id == meglevo.id).limit(1)) is not None
        if meglevo
        else False
    )
    # A KÉZZEL beállított oszlop-ember kötések pillanatképe (cimke ->
    # employee_id): a csere után visszaörököljük, ha az automatikus névre-
    # kötés nem talál egyértelmű embert - a felületen egyszer már rendbe tett
    # kötés ne vesszen el minden szinkronnál.
    korabbi_kotesek: dict[str, int] = {}
    # Ugyanígy a kézzel ELREJTETT oszlopok is a feliratukon át öröklődnek -
    # egy szinkron ne hozza vissza az egyszer már eltüntetett oszlopokat.
    korabbi_rejtettek: set[str] = set()
    # És a kézzel ELREJTETT SOROK is (a felhasznaló kérése: a szinkron ne
    # hozza vissza őket): dátumos sort a (dátum, diszpószám) párja azonosít
    # újra, dátum nélkülit a helye (idx) - a Sheet sorai jellemzően nem
    # csúsznak el, és ha mégis, az elrejtés visszaadható egy kattintással.
    korabbi_rejtett_sor_kulcsok: set[tuple[date, int | None]] = set()
    korabbi_rejtett_sor_idxek: set[int] = set()
    if meglevo is not None:
        for regi in db.scalars(select(DiszpoOszlop).where(DiszpoOszlop.munkalap_id == meglevo.id)).all():
            if regi.employee_id is not None and regi.cimke:
                korabbi_kotesek[ekezet_nelkul(regi.cimke.strip())] = regi.employee_id
            if regi.rejtett and regi.cimke:
                korabbi_rejtettek.add(ekezet_nelkul(regi.cimke.strip()))
        for regi_sor in db.scalars(select(DiszpoSor).where(DiszpoSor.munkalap_id == meglevo.id)).all():
            if not regi_sor.rejtett:
                continue
            if regi_sor.datum is not None:
                korabbi_rejtett_sor_kulcsok.add((regi_sor.datum, regi_sor.diszposzam))
            else:
                korabbi_rejtett_sor_idxek.add(regi_sor.idx)

    # A CELLÁK
    cellak: list[tuple[int, int, str | None, str | None]] = []
    for r in range(1, max_sor + 1):
        for c in range(1, max_oszlop + 1):
            cella = ws.cell(r, c)
            ertek = cella_erteke(cella.value)
            szin = cella_szine(cella)
            if ertek is None and szin is None:
                continue
            cellak.append((r - 1, c - 1, ertek, szin))

    # AZ OSZLOPOK: a nevek sora a címke (lásd nevsor_indexe), a belsősön a
    # legfelső sor a szekció - a szekció a Sheetben csak az első oszlopánál
    # áll ott, tehát jobbra "átfolyik".
    nevsor = nevsor_indexe(ws, fejlec_sorok, max_oszlop)
    terkep = belsos_nevterkep(db)
    oszlopok: list[dict] = []
    uzenetek: list[str] = []
    aktualis_csoport: str | None = None
    for c in range(1, max_oszlop + 1):
        if ketsoros:
            csoport_ertek = cella_erteke(ws.cell(1, c).value)
            if csoport_ertek:
                aktualis_csoport = csoport_ertek
        cimke = cella_erteke(ws.cell(nevsor, c).value)
        # Az automatikus névre-kötés csak a BELSŐS táblán fut: a külsős
        # oszlopfeliratok külsősök/eszközök, egy véletlen név-egyezés ott
        # rossz ember napjait számolná.
        ember, uzenet = oszlop_embere(cimke, terkep) if ketsoros else (None, None)
        employee_id = ember.id if ember else None
        if employee_id is None and cimke:
            # A korábbi, kézzel megadott kötés visszaöröklése.
            employee_id = korabbi_kotesek.get(ekezet_nelkul(cimke.strip()))
            if employee_id is not None:
                uzenet = None
        if uzenet:
            uzenetek.append(uzenet)
        oszlopok.append(
            {
                "idx": c - 1,
                "cimke": cimke,
                "csoport": aktualis_csoport if fejlec_sorok == 2 else None,
                "employee_id": employee_id,
                "rejtett": bool(cimke) and ekezet_nelkul(cimke.strip()) in korabbi_rejtettek,
            }
        )

    # A SOROK: a dátum TOVÁBBVITELE a lényeg - egy naphoz két diszpó is
    # tartozhat, és a második sor dátum-mezője üres.
    sorok: list[dict] = []
    utolso_datum: date | None = None
    for r in range(1, max_sor + 1):
        nyers_datum = ws.cell(r, 1).value if max_oszlop >= 1 else None
        if isinstance(nyers_datum, datetime):
            utolso_datum = nyers_datum.date()
        elif isinstance(nyers_datum, date):
            utolso_datum = nyers_datum
        ertekek = [cella_erteke(ws.cell(r, c).value) for c in range(1, min(max_oszlop, 8) + 1)]
        elvalaszto = elvalaszto_sor(ertekek)
        if elvalaszto:
            utolso_datum = None
        nap = cella_erteke(ws.cell(r, 2).value) if max_oszlop >= 2 else None
        diszposzam = cella_erteke(ws.cell(r, 3).value) if max_oszlop >= 3 else None
        sor_datum = utolso_datum if r > fejlec_sorok and not elvalaszto else None
        sor_diszposzam = int(diszposzam) if (diszposzam or "").isdigit() else None
        sorok.append(
            {
                "idx": r - 1,
                "datum": sor_datum,
                "nap": nap if isinstance(nap, str) and len(nap) < 20 else None,
                "diszposzam": sor_diszposzam,
                "elvalaszto": elvalaszto,
                # A kézzel elrejtett sor a csere után is rejtett marad - lásd
                # fent a korabbi_rejtett_sor_* gyűjtését.
                "rejtett": (
                    (sor_datum, sor_diszposzam) in korabbi_rejtett_sor_kulcsok
                    if sor_datum is not None
                    else (r - 1) in korabbi_rejtett_sor_idxek
                ),
            }
        )

    osszegzes = {
        "munkalap": ws.title,
        "sorok": len(sorok),
        "oszlopok": len(oszlopok),
        "cellak": len(cellak),
        "szines": sum(1 for *_, szin in cellak if szin),
        "emberhez_kotve": sum(1 for o in oszlopok if o["employee_id"]),
        "uzenetek": uzenetek,
        "felulir": regi_cellak,
    }
    if not vegrehajt:
        return osszegzes

    if meglevo is None:
        meglevo = DiszpoMunkalap(nev=ws.title)
        db.add(meglevo)
    meglevo.sorrend = sorrend
    meglevo.sor_szam = len(sorok)
    meglevo.oszlop_szam = len(oszlopok)
    meglevo.fejlec_sorok = fejlec_sorok
    db.flush()

    # CSERE, nem összefésülés: a Sheet az igazság ennél az átvételnél.
    for tabla in (DiszpoCella, DiszpoSor, DiszpoOszlop):
        db.query(tabla).filter(tabla.munkalap_id == meglevo.id).delete(synchronize_session=False)
    db.flush()

    db.bulk_save_objects([DiszpoOszlop(munkalap_id=meglevo.id, **o) for o in oszlopok])
    db.bulk_save_objects([DiszpoSor(munkalap_id=meglevo.id, **s) for s in sorok])
    db.bulk_save_objects(
        [
            DiszpoCella(munkalap_id=meglevo.id, sor_idx=r, oszlop_idx=c, ertek=e, szin=sz)
            for r, c, e, sz in cellak
        ]
    )
    db.commit()
    return osszegzes


def teljes_szinkron(
    db: Session,
    xlsx_adat: bytes | None = None,
    vegrehajt: bool = True,
    munkalapok: list[str] | None = None,
    tablazat_id: str = TABLAZAT_ID,
) -> list[dict]:
    """Az egész munkafüzet szinkronja - munkalaponkénti összegzésekkel tér
    vissza. `xlsx_adat` nélkül letölti a megosztott táblázatot."""
    if xlsx_adat is None:
        xlsx_adat = letoltes(tablazat_id)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_adat))
    eredmeny: list[dict] = []
    for sorrend, nev in enumerate(wb.sheetnames):
        if munkalapok and nev not in munkalapok:
            continue
        eredmeny.append(munkalap_atvetele(db, wb[nev], sorrend, vegrehajt))
    return eredmeny
