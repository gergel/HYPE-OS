"""A Krumpello kassza és munkabér szinkronja a Google Sheets táblázatból.

Forrás: a "HYPE PRODUCTIONS KFT. 2026 - PÉNZÜGY" munkafüzet
"KRUMPELLO - KASSZA" és "KRUMPELLO - MUNKABÉR" lapjai - ugyanaz a letöltési
mechanizmus, mint a HYPE 2026 diszpó-táblánál (lásd diszpo_sheet_sync.py):
a linkkel megosztott munkafüzet export?format=xlsx végpontja.

A feldolgozó logika a scripts/krumpello_import.py-ból költözött ide, hogy a
felületről indított háttér-szinkron (lásd routes/krumpello.py "sheet-sync")
és a parancssori script UGYANAZT a kódot futtassa - két külön másolat
előbb-utóbb mást importálna.

IDEMPOTENS, és a felületről indítva FELÜLÍRÓ: a táblázat az igazság (a
felhasználó kérése) - a napi kassza-sorok értékei a táblázat szerintire
állnak. Amit viszont a táblázat nem tartalmaz, azt nem töröljük: a rendszerben
felvitt, a táblázatba (még) fel nem vezetett tételek megmaradnak, és a napló
jelzi a darabszámukat.
"""

from __future__ import annotations

import datetime
import io
import re
from collections.abc import Callable

import openpyxl
import requests
from sqlalchemy.orm import Session

from app.models.krumpello import KrumpelloDolgozo, KrumpelloKiadas, KrumpelloMunkaora, KrumpelloNap

#: A megosztott munkafüzet azonosítója (a link /d/ és /edit közti része).
TABLAZAT_ID = "1NErnf6oixfTPq-J6v8BugwvcKh0P3-va8OSp4pszAsM"

KASSZA_LAP = "KRUMPELLO - KASSZA"
MUNKABER_LAP = "KRUMPELLO - MUNKABÉR"

#: A kiadás-blokkok a kassza-lapon: a "KIADÁS" alatti három szakasz. A
#: fejlécben szereplő nevük köti őket a modell `forras` értékeihez.
KIADAS_BLOKKOK = {
    "UTALÁS / BANKKÁRTYA": "utalas",
    "KÉSZPÉNZ": "keszpenz",
    "EXTRA": "extra",
}


class SheetLetoltesHiba(RuntimeError):
    """A munkafüzet nem tölthető le - jellemzően nincs linkkel megosztva."""


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def letoltes(db: Session | None = None, tablazat_id: str = TABLAZAT_ID) -> bytes:
    """A munkafüzet letöltése - ELŐSZÖR a bekötött Google-fiókkal.

    A pénzügy-táblázat nem nyilvános linkkel, hanem a felhasználó fiókjával
    van megosztva - a Beállításokban bekötött Google-fiók (lásd
    services/google_oauth.py) Drive-olvasó jogával exportáljuk. Ha nincs
    bekötve, vagy a token még a Drive-jog nélküli régi bekötésből való,
    a nyilvános export URL a tartalék."""
    if db is not None:
        try:
            from app.services import google_oauth

            creds = google_oauth.load_credentials(db)
            if creds is not None:
                valasz = requests.get(
                    f"https://www.googleapis.com/drive/v3/files/{tablazat_id}/export",
                    params={"mimeType": XLSX_MIME},
                    headers={"Authorization": f"Bearer {creds.token}"},
                    timeout=120,
                )
                if valasz.status_code == 200 and not valasz.content[:100].lstrip().startswith(
                    (b"<!DOCTYPE", b"<html")
                ):
                    return valasz.content
        except Exception:  # noqa: BLE001 - a hitelesített út hibája ne állítsa meg a tartalék utat
            pass

    cim = f"https://docs.google.com/spreadsheets/d/{tablazat_id}/export?format=xlsx"
    valasz = requests.get(cim, timeout=120)
    valasz.raise_for_status()
    # Ha a táblázat nincs megosztva és a fiókos út sem járt sikerrel, a Google
    # egy bejelentkezési HTML-oldalt ad vissza 200-zal - azt openpyxl-hibaként
    # elszállni hagyni csak egy rejtélyes "not a zip file"-t adna.
    if valasz.content[:100].lstrip().startswith((b"<!DOCTYPE", b"<html")):
        raise SheetLetoltesHiba(
            "A táblázat nem tölthető le. Két megoldás van: a Beállításokban kösd újra a "
            "Google-fiókot (az új bekötés már Drive-olvasó jogot is kap, és azzal a fiókkal "
            "megosztott táblázat is elérhető), VAGY oszd meg a munkafüzetet 'Bárki a "
            "linkkel: Megtekintő' módban."
        )
    return valasz.content


def _datum(ertek) -> datetime.date | None:
    if isinstance(ertek, datetime.datetime):
        return ertek.date()
    if isinstance(ertek, datetime.date):
        return ertek
    return None


def _szam(ertek) -> float | None:
    return float(ertek) if isinstance(ertek, (int, float)) and not isinstance(ertek, bool) else None


#: Szövegben álló összeg: "6535FT", "BÓNUSZ: 5000 FT", "1700 Ft - Tüdőszűrés".
_OSSZEG_MINTA = re.compile(r"-?\s*\d[\d\s .]*", re.UNICODE)

#: Ha ez szerepel a szövegben, az összeg egy IDŐSZAKI VÉGÖSSZEG, nem az adott
#: napé - napi értékként felvéve megduplázná az elszámolást.
_VEGOSSZEG_SZAVAK = ("összesen", "osszesen", "kiegészítés", "kiegeszites")


def _szam_szovegbol(ertek) -> tuple[float | None, str | None]:
    """(összeg, eredeti szöveg) egy cellából - lásd a script eredeti
    kommentjét: a borravaló oszlopba jegyzet is kerül, nem csak szám."""
    szam = _szam(ertek)
    if szam is not None:
        return szam, None
    szoveg = _szoveg(ertek)
    if szoveg is None:
        return None, None
    egysoros = " ".join(szoveg.split())
    if any(sz in egysoros.casefold() for sz in _VEGOSSZEG_SZAVAK):
        return None, egysoros
    talalat = _OSSZEG_MINTA.search(egysoros)
    if talalat is None:
        return None, egysoros
    nyers = talalat.group(0).replace(" ", "").replace(" ", "").replace(".", "")
    try:
        return float(nyers), egysoros
    except ValueError:
        return None, egysoros


def _szoveg(ertek) -> str | None:
    if ertek is None:
        return None
    szoveg = str(ertek).strip()
    return szoveg or None


def _blokk_oszlopai(ws, fejlec_sor: int, kezdet: int, veg: int) -> dict[str, int]:
    """Egy blokk oszlopnév -> oszlopindex térképe a fejlécsorból."""
    return {
        str(ws.cell(fejlec_sor, c).value).strip().upper(): c
        for c in range(kezdet, veg)
        if ws.cell(fejlec_sor, c).value
    }


def _kiadas_blokkok(ws, blokk_fejlec_sor: int, oszlop_fejlec_sor: int) -> list[tuple[str, dict[str, int]]]:
    """A három kiadás-blokk (forrás, oszloptérkép) párokként."""
    kezdetek: list[tuple[int, str]] = []
    for c in range(1, ws.max_column + 1):
        cimke = _szoveg(ws.cell(blokk_fejlec_sor, c).value)
        if cimke and cimke.upper() in KIADAS_BLOKKOK:
            kezdetek.append((c, KIADAS_BLOKKOK[cimke.upper()]))
    blokkok = []
    for i, (c, forras) in enumerate(kezdetek):
        veg = kezdetek[i + 1][0] if i + 1 < len(kezdetek) else ws.max_column + 1
        blokkok.append((forras, _blokk_oszlopai(ws, oszlop_fejlec_sor, c, veg)))
    return blokkok


def szinkron(db: Session, wb, *, felulir: bool = True, naplo: Callable[[str], None] = lambda s: None) -> dict:
    """A munkafüzet feldolgozása a megnyitott adatbázis-munkameneten.

    NEM commitol - a hívó dönti el (a script és a háttérfeladat is maga zár).
    Az összegzést dict-ben adja vissza; a részletek a naplóba mennek."""
    for lap in (KASSZA_LAP, MUNKABER_LAP):
        if lap not in wb.sheetnames:
            raise SheetLetoltesHiba(f"Hiányzik a(z) {lap!r} lap a munkafüzetből. Talált lapok: {wb.sheetnames}")

    uj_nap = uj_kiadas = uj_ora = uj_dolgozo = 0
    frissitett = 0

    # ── Napi kassza ────────────────────────────────────────────────────────
    ws = wb[KASSZA_LAP]
    bev = _blokk_oszlopai(ws, 3, 1, 13)
    csoportok: list[tuple[str, int]] = []
    for c in range(1, 13):
        cimke = _szoveg(ws.cell(2, c).value)
        if cimke:
            csoportok.append((cimke.upper(), c))
    kp_kartya = {}
    for cimke, c in csoportok:
        if cimke in ("BRUTTÓ", "NETTÓ", "BORRAVALÓ"):
            kp_kartya[cimke] = (c, c + 1)
    extra_oszlop = bev.get("EXTRA")

    letezo_napok = {n.datum: n for n in db.query(KrumpelloNap).all()}
    tablazat_napok: set[datetime.date] = set()
    for r in range(4, ws.max_row + 1):
        d = _datum(ws.cell(r, 1).value)
        if d is None:
            continue
        ertekek = {}
        for cimke, kulcs in (("BRUTTÓ", "brutto"), ("NETTÓ", "netto"), ("BORRAVALÓ", "borravalo")):
            if cimke in kp_kartya:
                kp, kartya = kp_kartya[cimke]
                ertekek[f"{kulcs}_kp"] = _szam(ws.cell(r, kp).value)
                ertekek[f"{kulcs}_kartya"] = _szam(ws.cell(r, kartya).value)
        ertekek["extra"] = _szam(ws.cell(r, extra_oszlop).value) if extra_oszlop else None
        # A csupa NULLÁS nap is bejön - az üres (előre kirajzolt) sorok nem.
        if all(v is None for v in ertekek.values()):
            continue
        tablazat_napok.add(d)
        meglevo = letezo_napok.get(d)
        if meglevo is None:
            db.add(KrumpelloNap(datum=d, **ertekek))
            uj_nap += 1
        elif felulir:
            valtozott = any(
                (getattr(meglevo, k) is None) != (v is None)
                or (v is not None and float(getattr(meglevo, k) or 0) != v)
                for k, v in ertekek.items()
            )
            for k, v in ertekek.items():
                setattr(meglevo, k, v)
            if valtozott:
                frissitett += 1

    # ── Kiadások ───────────────────────────────────────────────────────────
    letezo_kiadas = {
        (k.forras, k.kedvezmenyezett, k.datum, k.megnevezes, float(k.brutto) if k.brutto is not None else None)
        for k in db.query(KrumpelloKiadas).all()
    }
    tablazat_kiadasok: set = set()
    for forras, oszlopok in _kiadas_blokkok(ws, 2, 3):
        kedv = oszlopok.get("KEDVEZMÉNYEZETT")
        if kedv is None:
            continue
        for r in range(4, ws.max_row + 1):
            nev = _szoveg(ws.cell(r, kedv).value)
            if not nev:
                continue
            sor = {
                "forras": forras,
                "kedvezmenyezett": nev,
                "datum": _datum(ws.cell(r, oszlopok["DÁTUM"]).value) if "DÁTUM" in oszlopok else None,
                "megnevezes": _szoveg(ws.cell(r, oszlopok["TÉTEL NEVE"]).value) if "TÉTEL NEVE" in oszlopok else None,
                "netto": _szam(ws.cell(r, oszlopok["NETTÓ"]).value) if "NETTÓ" in oszlopok else None,
                "afa": _szam(ws.cell(r, oszlopok["ÁFA"]).value) if "ÁFA" in oszlopok else None,
            }
            brutto_oszlop = oszlopok.get("BRUTTÓ") or oszlopok.get("ÖSSZEG")
            sor["brutto"] = _szam(ws.cell(r, brutto_oszlop).value) if brutto_oszlop else None
            kulcs = (sor["forras"], sor["kedvezmenyezett"], sor["datum"], sor["megnevezes"], sor["brutto"])
            tablazat_kiadasok.add(kulcs)
            if kulcs in letezo_kiadas:
                continue
            letezo_kiadas.add(kulcs)
            db.add(KrumpelloKiadas(**sor))
            uj_kiadas += 1

    # ── Dolgozók és munkaóra ───────────────────────────────────────────────
    ws2 = wb[MUNKABER_LAP]
    emberek = [(str(ws2.cell(1, c).value).strip(), c) for c in range(1, ws2.max_column + 1) if ws2.cell(1, c).value]
    dolgozok = {d.nev.casefold(): d for d in db.query(KrumpelloDolgozo).all()}
    letezo_ora = {(o.dolgozo_id, o.datum) for o in db.query(KrumpelloMunkaora).all()}

    sor_datumok: dict[int, datetime.date] = {}
    zaras_sorok: dict[int, str] = {}
    for _, kezdo in emberek:
        for r in range(3, ws2.max_row + 1):
            nyers = ws2.cell(r, kezdo).value
            d = _datum(nyers)
            if d is not None:
                sor_datumok.setdefault(r, d)
                continue
            szoveg = _szoveg(nyers)
            if szoveg and "ZÁRÁS" in szoveg.upper():
                zaras_sorok.setdefault(r, " ".join(szoveg.split()))
    for r in range(3, ws2.max_row + 1):
        if r in sor_datumok or r in zaras_sorok:
            continue
        if any(_szam(ws2.cell(r, kezdo + 4).value) for _, kezdo in emberek):
            zaras_sorok[r] = f"{r}. sor (felirat nélküli időszaki zárás)"

    zarasok: list[str] = []
    datum_nelkul: list[str] = []

    for nyers_nev, kezdo in emberek:
        nev = nyers_nev.title() if nyers_nev.isupper() else nyers_nev
        dolgozo = dolgozok.get(nev.casefold())
        if dolgozo is None:
            dolgozo = KrumpelloDolgozo(nev=nev)
            db.add(dolgozo)
            db.flush()
            dolgozok[nev.casefold()] = dolgozo
            uj_dolgozo += 1
        oszlopok = _blokk_oszlopai(ws2, 2, kezdo, kezdo + 6)
        for r in range(3, ws2.max_row + 1):
            nyers_datum = ws2.cell(r, oszlopok["DÁTUM"]).value if "DÁTUM" in oszlopok else None
            ora = _szam(ws2.cell(r, oszlopok["ÓRA"]).value) if "ÓRA" in oszlopok else None
            orabar = _szam(ws2.cell(r, oszlopok["ÓRABÉR"]).value) if "ÓRABÉR" in oszlopok else None
            fizetes = _szam(ws2.cell(r, oszlopok["FIZETÉS"]).value) if "FIZETÉS" in oszlopok else None
            borravalo, jegyzet = (
                _szam_szovegbol(ws2.cell(r, oszlopok["BORRAVALÓ"]).value)
                if "BORRAVALÓ" in oszlopok
                else (None, None)
            )

            if r in zaras_sorok:
                if fizetes or borravalo or jegyzet:
                    zarasok.append(f"{nev}: {zaras_sorok[r]} → {fizetes or 0:,.0f} Ft")
                continue

            d = _datum(nyers_datum) or sor_datumok.get(r)
            if d is None:
                if ora or fizetes or borravalo:
                    datum_nelkul.append(f"{nev} ({ws2.title} {r}. sor)")
                continue

            if not ora and not borravalo and not fizetes and not jegyzet:
                continue
            if (dolgozo.id, d) in letezo_ora:
                continue
            letezo_ora.add((dolgozo.id, d))
            if fizetes is None and ora is not None and orabar is not None:
                fizetes = round(ora * orabar, 2)
            db.add(
                KrumpelloMunkaora(
                    dolgozo_id=dolgozo.id, datum=d, ora=ora, orabar=orabar,
                    fizetes=fizetes, borravalo=borravalo, megjegyzes=jegyzet,
                )
            )
            uj_ora += 1
            if orabar:
                dolgozo.alap_orabar = orabar

    # ── Napló és összegzés ─────────────────────────────────────────────────
    if zarasok:
        naplo("IDŐSZAKI ZÁRÁS-sorok (nem munkanapok, ezért NEM jönnek át - a napi soraik összege adja ki ugyanezt):")
        for z in zarasok:
            naplo(f"  {z}")
    if datum_nelkul:
        naplo("FIGYELEM - ezekhez nem volt dátum, és a sorból sem volt pótolható:")
        for x in datum_nelkul:
            naplo(f"  {x}")
    # Amit a rendszer tartalmaz, de a táblázat nem - ezeket NEM töröljük, csak
    # jelezzük: lehetnek itt felvitt, a táblázatba még fel nem vezetett sorok.
    csak_helyi_nap = sorted(set(letezo_napok) - tablazat_napok)
    if csak_helyi_nap:
        naplo(
            f"MEGJEGYZÉS: {len(csak_helyi_nap)} kassza-nap csak a rendszerben van meg, a táblázatban nincs "
            f"(nem töröltük): {', '.join(str(d) for d in csak_helyi_nap[:20])}"
            + ("…" if len(csak_helyi_nap) > 20 else "")
        )

    naplo(f"Új nap: {uj_nap}" + (f" (felülírva/frissítve: {frissitett})" if felulir else ""))
    naplo(f"Új kiadás: {uj_kiadas}")
    naplo(f"Új dolgozó: {uj_dolgozo}")
    naplo(f"Új munkaóra: {uj_ora}")

    return {
        "uj_nap": uj_nap,
        "frissitett_nap": frissitett,
        "uj_kiadas": uj_kiadas,
        "uj_dolgozo": uj_dolgozo,
        "uj_munkaora": uj_ora,
        "uzenet": (
            f"Új nap: {uj_nap} (frissítve: {frissitett}) · új kiadás: {uj_kiadas} · "
            f"új dolgozó: {uj_dolgozo} · új munkaóra: {uj_ora}"
        ),
    }


def teljes_szinkron(
    db: Session,
    *,
    xlsx_adat: bytes | None = None,
    felulir: bool = True,
    naplo: Callable[[str], None] = lambda s: None,
) -> dict:
    """Letöltés + feldolgozás egyben - a felületi háttér-szinkron ezt hívja.
    `xlsx_adat`-tal a letöltés kihagyható (teszt / kézi fájl)."""
    if xlsx_adat is None:
        xlsx_adat = letoltes(db)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_adat), data_only=True)
    return szinkron(db, wb, felulir=felulir, naplo=naplo)
